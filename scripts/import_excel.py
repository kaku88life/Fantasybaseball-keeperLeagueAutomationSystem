"""
Fantasy Baseball Keeper League - Excel Roster Importer

Parses the 5-Man Keep盟 roster Excel file.

Excel Structure:
  - Yearly sheets: "2023年選秀前名單", "2024年選秀前名單", "2025年選秀前名單", "2026年選秀前名單"
  - Individual team sheets: "楊善合", "TIMMY LIU", "Leo", etc. (16 teams)

Layout (yearly sheets):
  - 4 teams per horizontal block, each block ~28 rows
  - Block 1: cols B-D, G-I, L-N, Q-S (teams 1-4)
  - Block 2: same column pattern, offset by ~28 rows (teams 5-8)
  - Block 3: teams 9-12
  - Block 4: teams 13-16

Each team block:
  - Header row: "2025季初" + team Yahoo name
  - Subheader: manager name + "合約金額$" + "合約型態"
  - Player rows: position | player_name | salary (number) | contract_type (string)
  - "Total keeper cost:" row
  - Financial summary rows: 起始資金, 前季排名獎勵金, 交易補償金, 買斷金, FAAB
  - Buyout section: 買斷 header + buyout records

Formats:
  - 2023 sheet: contract as single string "$16/N3" in one column
  - 2024+ sheets: contract split into salary (number) + type (string) columns
"""
from __future__ import annotations

import sys
from pathlib import Path

# Add project root to path
project_root = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(project_root))

import openpyxl

from config.settings import FAAB_BASE, get_salary_cap
from src.contract.models import (
    BuyoutRecord,
    Contract,
    ContractType,
    Player,
    Team,
)
from src.parser.normalizer import (
    detect_special_status,
    normalize_player_name,
    normalize_position,
    parse_buyout_string,
    parse_contract_columns,
    parse_contract_string,
)


# Column offsets for the 4-team-per-row layout (2024+ sheets)
# Each team group: [position_col, name_col, salary_col, contract_col, note_col]
TEAM_COLUMN_GROUPS = [
    {"pos": 1, "name": 2, "salary": 3, "contract": 4, "note": 5},      # Team A: cols A-E
    {"pos": 6, "name": 7, "salary": 8, "contract": 9, "note": 10},     # Team B: cols F-J
    {"pos": 11, "name": 12, "salary": 13, "contract": 14, "note": 15}, # Team C: cols K-O
    {"pos": 16, "name": 17, "salary": 18, "contract": 19, "note": 20}, # Team D: cols P-T
]

# Buyout section column groups
BUYOUT_COLUMN_GROUPS = [
    {"label": 1, "name": 2, "detail": 3, "faab": 4},        # Team A
    {"label": 6, "name": 7, "detail": 8, "faab": 9},        # Team B
    {"label": 11, "name": 12, "detail": 13, "faab": 14},     # Team C
    {"label": 16, "name": 17, "detail": 18, "faab": 19},     # Team D
]

# Financial row labels
FINANCIAL_LABELS = {
    "Total keeper cost:": "keeper_cost",
    "起始資金": "salary_cap",
    "2025起始資金": "salary_cap",
    "2024起始資金": "salary_cap",
    "2026起始資金": "salary_cap",
    "前季排名獎勵金": "ranking_bonus",
    "交易補償金": "trade_compensation",
    "買斷金": "buyout_cost",
    "FAAB": "faab",
}


def load_workbook(filepath: str) -> openpyxl.Workbook:
    """Load Excel workbook."""
    return openpyxl.load_workbook(filepath, data_only=True)


def get_cell(ws, row: int, col: int):
    """Get cell value safely."""
    val = ws.cell(row=row, column=col).value
    return val


def import_yearly_sheet(ws, year: int) -> list[Team]:
    """
    Import a yearly roster sheet (e.g., "2025年選秀前名單").

    The sheet has 4 horizontal blocks, each containing 4 teams.
    Total: 16 teams.

    Two layouts:
    - 2023: no "合約金額$" header, contract as single string, cols (1-3, 5-7, 9-11, 13-15)
    - 2024+: has "合約金額$" header, salary + type split, cols (1-4, 6-9, 11-14, 16-19)
    """
    # Detect format: check if any cell contains "合約金額"
    has_salary_header = False
    for row_idx in range(1, min(10, ws.max_row + 1)):
        for col_idx in range(1, 25):
            val = get_cell(ws, row_idx, col_idx)
            if val and isinstance(val, str) and "合約金額" in str(val):
                has_salary_header = True
                break
        if has_salary_header:
            break

    if has_salary_header:
        # 2024+ format
        team_blocks = _find_team_blocks(ws)
        teams = []
        for block in team_blocks:
            team = _parse_team_block(ws, block, year)
            if team:
                teams.append(team)
        return teams
    else:
        # 2023 format
        return _import_2023_sheet(ws, year)


def _find_team_blocks(ws) -> list[dict]:
    """
    Find all team blocks in the worksheet.
    A team block starts with a header row containing "季初" and ends before the next block.

    Returns list of dicts: {
        "header_row": int,      # row with "2025季初" + team name
        "subheader_row": int,   # row with manager name + "合約金額$"
        "col_group": dict,      # column group from TEAM_COLUMN_GROUPS
        "buyout_col_group": dict,
        "manager_name": str,
        "team_name": str,
    }
    """
    blocks = []
    max_row = ws.max_row

    for row_idx in range(1, max_row + 1):
        for group_idx, col_group in enumerate(TEAM_COLUMN_GROUPS):
            name_col = col_group["name"]
            salary_col = col_group["salary"]

            cell_val = get_cell(ws, row_idx, name_col)
            salary_val = get_cell(ws, row_idx, salary_col)

            if cell_val and isinstance(cell_val, str) and "季初" in str(cell_val):
                # This is a header row: "2025季初" is in name col, team Yahoo name is in salary col
                # But need to check: sometimes the pattern differs
                pass

            # Look for subheader pattern: manager name + "合約金額$"
            if salary_val and isinstance(salary_val, str) and "合約金額" in str(salary_val):
                manager_name = str(cell_val).strip() if cell_val else ""
                # Get team name from the row above
                team_name = ""
                header_val = get_cell(ws, row_idx - 1, salary_col)
                if header_val and isinstance(header_val, str):
                    team_name = str(header_val).strip()

                blocks.append({
                    "header_row": row_idx - 1,
                    "subheader_row": row_idx,
                    "col_group": col_group,
                    "buyout_col_group": BUYOUT_COLUMN_GROUPS[group_idx],
                    "manager_name": manager_name,
                    "team_name": team_name,
                })

    return blocks


# 2023 format column groups: (pos_col, name_col, contract_col)
# Layout: cols 1-3, 5-7, 9-11, 13-15 (4 teams per row, tighter spacing)
TEAM_COLUMN_GROUPS_2023 = [
    {"pos": 1, "name": 2, "contract": 3},   # Team A
    {"pos": 5, "name": 6, "contract": 7},   # Team B
    {"pos": 9, "name": 10, "contract": 11},  # Team C
    {"pos": 13, "name": 14, "contract": 15}, # Team D
]


def _import_2023_sheet(ws, year: int) -> list[Team]:
    """
    Import 2023-format yearly sheet.

    2023 layout:
    - Row 1: "2023季初" headers in name columns (cols 2, 6, 10, 14)
    - Row 2: manager names in name columns
    - Row 3+: position | player_name | contract_string (e.g., "$16/N3")
    - End markers: "Total keeper cost:" row
    - Buyout section below
    - Then next 4 teams in rows below
    """
    teams = []

    # Find all blocks: look for rows where name columns have manager-like names
    # preceded by a "季初" row
    block_starts = []
    for row_idx in range(1, ws.max_row + 1):
        val = get_cell(ws, row_idx, 2)  # col B (first team's name col)
        if val and isinstance(val, str) and "季初" in str(val):
            block_starts.append(row_idx)

    for block_start in block_starts:
        manager_row = block_start + 1  # manager names in next row

        for group in TEAM_COLUMN_GROUPS_2023:
            name_col = group["name"]
            pos_col = group["pos"]
            contract_col = group["contract"]

            manager_name = get_cell(ws, manager_row, name_col)
            if not manager_name or not isinstance(manager_name, str):
                continue
            manager_name = str(manager_name).strip()
            if not manager_name:
                continue

            team = Team(
                manager_name=manager_name,
                team_name="",
            )

            # Parse player rows starting from manager_row + 1
            for row_idx in range(manager_row + 1, manager_row + 20):
                name_val = get_cell(ws, row_idx, name_col)
                if not name_val:
                    continue

                name_str = str(name_val).strip()
                if "Total keeper cost" in name_str:
                    cost_val = get_cell(ws, row_idx, contract_col)
                    if cost_val is not None:
                        try:
                            team._parsed_keeper_cost = int(float(cost_val))
                        except (ValueError, TypeError):
                            pass
                    break
                if name_str in ("", "買斷", "TOTAL"):
                    continue

                position_val = get_cell(ws, row_idx, pos_col)
                contract_val = get_cell(ws, row_idx, contract_col)

                position = normalize_position(str(position_val)) if position_val else ""
                player_name = normalize_player_name(name_str)

                # Parse contract string
                contract = None
                if contract_val:
                    contract = parse_contract_string(str(contract_val))

                if contract is None:
                    continue

                is_active = contract.contract_type != ContractType.R
                player = Player(
                    name=player_name,
                    position=position,
                    contract=contract,
                    is_active_keeper=is_active,
                )
                team.players.append(player)

            # Parse financial rows and buyout section below
            _parse_2023_financial_and_buyouts(ws, manager_row, group, team, year)

            teams.append(team)

    return teams


def _parse_2023_financial_and_buyouts(ws, manager_row: int, col_group: dict, team: Team, year: int):
    """Parse financial summary and buyout records for 2023 format."""
    name_col = col_group["name"]
    contract_col = col_group["contract"]
    pos_col = col_group["pos"]

    buyout_parsed = False  # prevent duplicate buyout parsing

    # Scan for financial labels and buyout sections
    for row_idx in range(manager_row + 1, manager_row + 30):
        pos_val = get_cell(ws, row_idx, pos_col)
        name_val = get_cell(ws, row_idx, name_col)
        label_str = str(pos_val).strip() if pos_val else ""
        name_str = str(name_val).strip() if name_val else ""

        # Financial labels (can be in pos_col or name_col)
        for check_str in (label_str, name_str):
            for pattern, field_name in FINANCIAL_LABELS.items():
                if pattern in check_str:
                    value = get_cell(ws, row_idx, contract_col)
                    try:
                        num_val = int(float(value)) if value is not None else 0
                    except (ValueError, TypeError):
                        num_val = 0

                    if field_name == "salary_cap":
                        team.salary_cap = num_val
                    elif field_name == "ranking_bonus":
                        team.ranking_bonus = num_val
                    elif field_name == "faab":
                        team.faab_budget = num_val
                    break

        # Buyout entries: exact "買斷" label (not "買斷金")
        if not buyout_parsed and label_str == "買斷":
            _collect_2023_buyout_rows(ws, row_idx, name_col, contract_col, team)
            buyout_parsed = True

    # Fallback: if no explicit "買斷" label found, scan for buyout strings
    # after financial rows (some teams list buyouts without a header)
    if not buyout_parsed:
        for row_idx in range(manager_row + 1, manager_row + 30):
            name_val = get_cell(ws, row_idx, name_col)
            contract_val = get_cell(ws, row_idx, contract_col)
            if not name_val or not contract_val:
                continue
            detail_str = str(contract_val).strip()
            # Buyout pattern: "$X/Y-Z=W" (has both "-" and "=")
            if "-" in detail_str and "=" in detail_str:
                parsed = parse_buyout_string(detail_str)
                if parsed:
                    b_name_str = str(name_val).strip()
                    note_val = get_cell(ws, row_idx, contract_col + 1)
                    note = str(note_val).strip() if note_val else ""
                    record = BuyoutRecord(
                        player_name=normalize_player_name(b_name_str),
                        original_contract=parsed["original_contract"],
                        buyout_salary_cost=parsed["salary_cost"],
                        buyout_faab_cost=parsed["faab_cost"],
                        remaining_years=1,
                        use_faab=parsed["faab_cost"] > 0,
                        note=note,
                    )
                    team.buyout_records.append(record)

    # Defaults
    if team.salary_cap == 0:
        team.salary_cap = get_salary_cap(year)
    if team.faab_budget == 0:
        team.faab_budget = FAAB_BASE


def _collect_2023_buyout_rows(ws, start_row: int, name_col: int, contract_col: int, team: Team):
    """Collect buyout records starting from a '買斷' label row."""
    for b_off in range(0, 5):
        b_row = start_row + b_off
        b_name = get_cell(ws, b_row, name_col)
        b_detail = get_cell(ws, b_row, contract_col)

        if not b_name:
            continue
        b_name_str = str(b_name).strip()
        if not b_name_str or "Total" in b_name_str:
            break

        parsed = parse_buyout_string(str(b_detail)) if b_detail else None
        if parsed:
            note_val = get_cell(ws, b_row, contract_col + 1)
            note = str(note_val).strip() if note_val else ""
            record = BuyoutRecord(
                player_name=normalize_player_name(b_name_str),
                original_contract=parsed["original_contract"],
                buyout_salary_cost=parsed["salary_cost"],
                buyout_faab_cost=parsed["faab_cost"],
                remaining_years=1,
                use_faab=parsed["faab_cost"] > 0,
                note=note,
            )
            team.buyout_records.append(record)


def _parse_team_block(ws, block: dict, year: int) -> Team | None:
    """Parse a single team block from the worksheet."""
    col = block["col_group"]
    buyout_col = block["buyout_col_group"]
    start_row = block["subheader_row"] + 1  # first player row

    team = Team(
        manager_name=block["manager_name"],
        team_name=block["team_name"],
    )

    # Parse player rows until we hit "Total keeper cost:" or empty section
    row_idx = start_row
    max_scan = 20  # max players per team

    for _ in range(max_scan):
        name_val = get_cell(ws, row_idx, col["name"])

        if not name_val:
            row_idx += 1
            continue

        name_str = str(name_val).strip()

        # Check for end markers
        if "Total keeper cost" in name_str:
            # Read keeper cost
            cost_val = get_cell(ws, row_idx, col["salary"])
            if cost_val is not None:
                team._parsed_keeper_cost = int(float(cost_val))
            break

        if name_str in ("", "買斷", "TOTAL"):
            row_idx += 1
            continue

        # Parse player
        position_val = get_cell(ws, row_idx, col["pos"])
        salary_val = get_cell(ws, row_idx, col["salary"])
        contract_val = get_cell(ws, row_idx, col["contract"])
        note_val = get_cell(ws, row_idx, col.get("note", 0))

        position = normalize_position(str(position_val)) if position_val else ""
        player_name = normalize_player_name(name_str)

        # Parse contract (handle both formats)
        contract = None
        if isinstance(salary_val, (int, float)) and contract_val:
            # 2024+ format: separate columns
            contract = parse_contract_columns(salary_val, str(contract_val))
        elif isinstance(salary_val, str) and "/" in str(salary_val):
            # 2023 format: single string in salary column
            contract = parse_contract_string(str(salary_val))
        elif contract_val and isinstance(contract_val, str) and "/" in str(contract_val):
            # Another 2023 variant: contract string in contract column
            contract = parse_contract_string(str(contract_val))

        if contract is None and salary_val is not None:
            # Fallback: try to extract from whatever we have
            try:
                salary = int(float(salary_val))
                ct_str = str(contract_val).strip().upper() if contract_val else "A"
                contract = parse_contract_columns(salary, ct_str)
            except (ValueError, TypeError):
                pass

        if contract is None:
            row_idx += 1
            continue

        # Check for special status
        special = detect_special_status(str(note_val)) if note_val else None
        if special and special.value != "none":
            contract.special_status = special

        # Determine if R contract -> bench keeper
        is_active = contract.contract_type != ContractType.R

        # Check note for "轉B" (R converted to B)
        if note_val and "轉B" in str(note_val):
            contract = Contract(
                contract_type=ContractType.B,
                salary=contract.salary,
            )
            is_active = True

        player = Player(
            name=player_name,
            position=position,
            contract=contract,
            is_active_keeper=is_active,
        )
        team.players.append(player)
        row_idx += 1

    # Parse financial summary rows (after Total keeper cost)
    _parse_financial_rows(ws, row_idx, col, team, year)

    # Parse buyout section
    _parse_buyout_section(ws, row_idx, buyout_col, team)

    return team


def _parse_financial_rows(ws, start_row: int, col: dict, team: Team, year: int):
    """Parse financial summary rows below the player list."""
    for offset in range(10):  # scan up to 10 rows below
        row_idx = start_row + offset
        label_val = get_cell(ws, row_idx, col["name"])

        if not label_val:
            continue

        label_str = str(label_val).strip()
        value = get_cell(ws, row_idx, col["salary"])

        # Match against known labels
        for pattern, field_name in FINANCIAL_LABELS.items():
            if pattern in label_str:
                try:
                    num_val = int(float(value)) if value is not None else 0
                except (ValueError, TypeError):
                    num_val = 0

                if field_name == "salary_cap":
                    team.salary_cap = num_val
                elif field_name == "ranking_bonus":
                    team.ranking_bonus = num_val
                elif field_name == "trade_compensation":
                    team.trade_compensation = num_val
                elif field_name == "faab":
                    team.faab_budget = num_val
                break

    # If salary_cap wasn't found from sheet, calculate from rules
    if team.salary_cap == 0:
        team.salary_cap = get_salary_cap(year)
    if team.faab_budget == 0:
        team.faab_budget = FAAB_BASE


def _parse_buyout_section(ws, start_row: int, buyout_col: dict, team: Team):
    """Parse buyout records below the financial summary."""
    # Find buyout header row (look for "買斷" label)
    for offset in range(15):
        row_idx = start_row + offset
        label_val = get_cell(ws, row_idx, buyout_col["label"])

        if label_val and "買斷" in str(label_val):
            # Parse buyout entries in subsequent rows
            for b_offset in range(1, 10):
                b_row = row_idx + b_offset
                b_name = get_cell(ws, b_row, buyout_col["name"])

                if not b_name:
                    # Check for "TOTAL" to stop
                    total_check = get_cell(ws, b_row, buyout_col["label"])
                    if total_check and "TOTAL" in str(total_check):
                        break
                    continue

                b_detail = get_cell(ws, b_row, buyout_col["detail"])
                b_faab = get_cell(ws, b_row, buyout_col["faab"])

                player_name = normalize_player_name(str(b_name))
                detail_str = str(b_detail) if b_detail else ""
                note_val = get_cell(ws, b_row, buyout_col.get("note", buyout_col["faab"] + 1))
                note = str(note_val).strip() if note_val else ""

                # Parse buyout detail
                parsed = parse_buyout_string(detail_str)

                if parsed:
                    record = BuyoutRecord(
                        player_name=player_name,
                        original_contract=parsed["original_contract"],
                        buyout_salary_cost=parsed["salary_cost"],
                        buyout_faab_cost=parsed["faab_cost"],
                        remaining_years=1,
                        use_faab=parsed["faab_cost"] > 0,
                        note=note,
                    )
                    team.buyout_records.append(record)
            break


def import_individual_sheet(ws) -> dict:
    """
    Import an individual team sheet (e.g., "楊善合").

    Structure:
    - Row 1: 球隊 | team Yahoo name
    - Row 2: 管理員 | manager name
    - Row 3: LINE | LINE name
    - Row 5+: seasonal columns (2023季初, 2024季初, 2025季初)
    - Each season column has player rows + buyout section

    Returns dict with multi-year roster history.
    """
    result = {
        "team_name": "",
        "manager_name": "",
        "line_name": "",
        "seasons": {},
    }

    # Parse header
    result["team_name"] = str(get_cell(ws, 1, 2) or "").strip()
    result["manager_name"] = str(get_cell(ws, 2, 2) or "").strip()
    result["line_name"] = str(get_cell(ws, 3, 2) or "").strip()

    # Find season column groups by looking for "季初" in row 5
    season_cols = {}
    for col_idx in range(1, ws.max_column + 1):
        val = get_cell(ws, 5, col_idx)
        if val and "季初" in str(val):
            year_str = str(val).strip()
            # Extract year: "2025季初" -> 2025
            import re
            year_match = re.search(r"(\d{4})", year_str)
            if year_match:
                year = int(year_match.group(1))
                season_cols[year] = col_idx

    # Parse each season
    for year, start_col in season_cols.items():
        players = []
        buyouts = []

        # Determine format based on year
        is_string_format = (year <= 2023)

        # Find the subheader row (manager name row)
        subheader_row = 6  # typically row 6

        # Parse players
        for row_idx in range(subheader_row + 1, ws.max_row + 1):
            # Check for end markers
            name_val = get_cell(ws, row_idx, start_col)
            if not name_val:
                continue

            name_str = str(name_val).strip()
            if "Total keeper cost" in name_str:
                break
            if name_str in ("買斷", "TOTAL", ""):
                # Check if buyout section
                if name_str == "買斷" or (get_cell(ws, row_idx, start_col - 1) and "買斷" in str(get_cell(ws, row_idx, start_col - 1))):
                    # Parse buyout rows below
                    for b_off in range(1, 10):
                        b_name = get_cell(ws, row_idx + b_off, start_col)
                        if not b_name:
                            break
                        b_detail = get_cell(ws, row_idx + b_off, start_col + 1)
                        parsed = parse_buyout_string(str(b_detail)) if b_detail else None
                        if parsed:
                            buyouts.append({
                                "player": normalize_player_name(str(b_name)),
                                **parsed,
                            })
                    break
                continue

            # Parse player row
            position_val = get_cell(ws, row_idx, start_col - 1)
            position = normalize_position(str(position_val)) if position_val else ""

            if is_string_format:
                # 2023: contract as string in column after name
                contract_str_val = get_cell(ws, row_idx, start_col + 1)
                contract = parse_contract_string(str(contract_str_val)) if contract_str_val else None
            else:
                # 2024+: salary + contract type in separate columns
                salary_val = get_cell(ws, row_idx, start_col + 1)
                contract_type_val = get_cell(ws, row_idx, start_col + 2)
                contract = parse_contract_columns(salary_val, str(contract_type_val)) if salary_val else None

            if contract:
                players.append({
                    "name": normalize_player_name(name_str),
                    "position": position,
                    "contract": contract.display,
                    "salary": contract.salary,
                    "contract_type": contract.contract_type.value,
                    "extension_years": contract.extension_years,
                })

        result["seasons"][year] = {
            "players": players,
            "buyouts": buyouts,
        }

    return result


def import_all(filepath: str) -> dict:
    """
    Import all data from the Excel file.

    Returns:
        {
            "yearly_rosters": {2023: [Team, ...], 2024: [...], ...},
            "individual_histories": {"楊善合": {...}, ...},
            "right_panel_finances": {manager_name: {salary_cap, faab}, ...},
        }
    """
    wb = load_workbook(filepath)
    result = {
        "yearly_rosters": {},
        "individual_histories": {},
        "right_panel_finances": {},
    }

    yearly_sheet_pattern = "年選秀前名單"
    individual_sheets = []

    for sheet_name in wb.sheetnames:
        if yearly_sheet_pattern in sheet_name:
            # Extract year
            import re
            year_match = re.search(r"(\d{4})", sheet_name)
            if year_match:
                year = int(year_match.group(1))
                print(f"Importing yearly sheet: {sheet_name} (year={year})")
                teams = import_yearly_sheet(wb[sheet_name], year)
                result["yearly_rosters"][year] = teams
                print(f"  Found {len(teams)} teams")
                for t in teams:
                    print(f"    {t.manager_name}: {len(t.players)} players, "
                          f"keeper cost=${t.total_keeper_cost}")
        else:
            individual_sheets.append(sheet_name)

    # Parse right-panel finance table from latest yearly sheet
    latest_year = max(result["yearly_rosters"].keys()) if result["yearly_rosters"] else None
    if latest_year:
        ws = wb[f"{latest_year}年選秀前名單"]
        result["right_panel_finances"] = _parse_right_panel_finances(ws)

    # Import individual team sheets
    for sheet_name in individual_sheets:
        print(f"Importing individual sheet: {sheet_name}")
        history = import_individual_sheet(wb[sheet_name])
        result["individual_histories"][sheet_name] = history

    return result


def _parse_right_panel_finances(ws) -> dict:
    """
    Parse the right-side finance summary table (cols V-X area).
    Contains each team's salary cap and FAAB after keeper selections.
    """
    finances = {}

    # The right panel starts around column 22 (V)
    # Look for header row with "起始資金" and "FAAB"
    for row_idx in range(1, 30):
        val_v = get_cell(ws, row_idx, 23)  # col W
        val_x = get_cell(ws, row_idx, 24)  # col X

        if val_v and "起始資金" in str(val_v):
            # Header found, parse rows below
            for data_row in range(row_idx + 1, row_idx + 20):
                manager = get_cell(ws, data_row, 22)  # col V
                salary = get_cell(ws, data_row, 23)    # col W
                faab = get_cell(ws, data_row, 24)      # col X

                if not manager:
                    break

                manager_str = str(manager).strip()
                try:
                    finances[manager_str] = {
                        "salary_cap": int(float(salary)) if salary else 0,
                        "faab": int(float(faab)) if faab else 0,
                    }
                except (ValueError, TypeError):
                    pass
            break

    return finances


# ========== CLI Entry Point ==========
if __name__ == "__main__":
    import json
    import argparse

    parser = argparse.ArgumentParser(description="Import Fantasy Baseball Excel roster data")
    parser.add_argument("filepath", help="Path to the Excel file")
    parser.add_argument("--year", type=int, help="Import only a specific year")
    parser.add_argument("--team", help="Import only a specific team (individual sheet)")
    parser.add_argument("--json", action="store_true", help="Output as JSON")
    args = parser.parse_args()

    data = import_all(args.filepath)

    if args.json:
        # Serialize for JSON output
        output = {
            "yearly_rosters": {},
            "individual_histories": data["individual_histories"],
            "right_panel_finances": data["right_panel_finances"],
        }
        for year, teams in data["yearly_rosters"].items():
            output["yearly_rosters"][year] = []
            for team in teams:
                output["yearly_rosters"][year].append({
                    "manager": team.manager_name,
                    "team_name": team.team_name,
                    "salary_cap": team.salary_cap,
                    "faab": team.faab_budget,
                    "ranking_bonus": team.ranking_bonus,
                    "trade_compensation": team.trade_compensation,
                    "players": [
                        {
                            "name": p.name,
                            "position": p.position,
                            "contract": p.contract.display,
                            "salary": p.contract.salary,
                        }
                        for p in team.players
                    ],
                    "buyouts": [
                        {
                            "player": b.player_name,
                            "original_contract": b.original_contract,
                            "salary_cost": b.buyout_salary_cost,
                            "faab_cost": b.buyout_faab_cost,
                        }
                        for b in team.buyout_records
                    ],
                })
        print(json.dumps(output, ensure_ascii=False, indent=2))
    else:
        # Summary output
        print("\n" + "=" * 60)
        print("Fantasy Baseball Keeper League - Excel Import Summary")
        print("=" * 60)

        for year in sorted(data["yearly_rosters"].keys()):
            teams = data["yearly_rosters"][year]
            print(f"\n--- {year} Season ({len(teams)} teams) ---")
            for team in teams:
                active = len(team.active_keepers)
                bench = len(team.bench_keepers)
                print(f"  {team.manager_name:20s} | "
                      f"Active: {active:2d} | "
                      f"Bench(R): {bench} | "
                      f"Cost: ${team.total_keeper_cost:3d} | "
                      f"Cap: ${team.salary_cap} | "
                      f"FAAB: ${team.faab_budget}")
                if team.buyout_records:
                    for b in team.buyout_records:
                        print(f"    Buyout: {b.player_name} {b.original_contract} "
                              f"(salary: ${b.buyout_salary_cost}, FAAB: ${b.buyout_faab_cost})")

        if data["right_panel_finances"]:
            print(f"\n--- Finance Summary (Right Panel) ---")
            for mgr, fin in data["right_panel_finances"].items():
                print(f"  {mgr:20s} | Cap: ${fin['salary_cap']:3d} | FAAB: ${fin['faab']}")
