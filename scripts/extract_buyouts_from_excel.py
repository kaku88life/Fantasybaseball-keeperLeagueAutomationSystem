"""Extract all buyout records from 2025 pre-draft Excel sheet."""
import openpyxl
import re
import json

EXCEL_PATH = r"C:\Users\amy41\OneDrive\Desktop\Baseball\5-Man Keep盟 新版球員名單.xlsx"
SHEET_NAME = "2025年選秀前名單"


def parse_buyout_string(buyout_str):
    """Parse buyout notation like '$11/O-6=5' or '16/O-8=8'."""
    if not buyout_str or not isinstance(buyout_str, str):
        return None
    buyout_str = buyout_str.strip()
    pattern = r"\$?(\d+)\s*/\s*([A-Za-z]\d*)\s*-\s*(\d+)\s*=\s*(\d+)"
    match = re.match(pattern, buyout_str)
    if not match:
        return None
    return {
        "salary": int(match.group(1)),
        "contract": match.group(2).upper(),
        "original_contract": f"${match.group(1)}/{match.group(2).upper()}",
        "faab_cost": int(match.group(3)),
        "salary_cost": int(match.group(4)),
    }


wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb[SHEET_NAME]

# Scan the entire sheet for buyout records
# Buyout sections start with a row containing "買斷" in col 1, 6, 11, or 16
# followed by rows with player names and buyout details

# Team layout: 4 teams per row block
# Block 1 (rows 5-32):  col 1-4 (楊善合), col 6-9 (TIMMY LIU), col 11-14 (林剛), col 16-19 (Hank)
# Block 2 (rows 33-61): col 1-4 (Leo), col 6-9 (Eddie), col 11-14 (Yu-Che), col 16-19 (Billy)
# Block 3 (rows 62-90): col 1-4 (James/魚魚), col 6-9 (Chih-Wei/wei), col 11-14 (Ponpon), col 16-19 (郭子睿/Kaku)
# Block 4 (rows 91-119): col 1-4 (Javier/謙謙), col 6-9 (Issac/rawstuff), col 11-14 (Tony), col 16-19 (ywchiou/YWC)

# Map Excel manager names to system names
MANAGER_MAP = {
    "楊善合": "哈寶好",
    "TIMMY LIU": "TIMMY LIU",
    "林剛": "Hyper",
    "Hank": "叫我寬哥",
    "Leo": "Leo",
    "Eddie Chen": "EDDIE",
    "Yu-Che Chang": "小喆",
    "Billy WU": "Billy",
    "James Chen": "魚魚",
    "Chih-Wei": "wei",
    "Ponpon": "Ponpon",
    "郭子睿(Rangers)": "Ｋａｋｕ",
    "Javier": "謙謙",
    "Issac": "rawstuff",
    "Tony林芳民": "Tony",
    "ywchiou": "YWC",
}

# Define the blocks: (manager_row, buyout_header_row, team_col_offsets)
# Each block has 4 teams at specific column positions
# We scan for "買斷" label rows and extract records below them

buyout_records = []

def get_cell(row, col):
    v = ws.cell(row=row, column=col).value
    return v

# Find all rows containing "買斷" label (buyout section headers)
for row_idx in range(1, ws.max_row + 1):
    for base_col in [1, 6, 11, 16]:
        cell_val = get_cell(row_idx, base_col)
        if cell_val and str(cell_val).strip() == "買斷":
            # This is a buyout section header
            # The manager name is found a few rows above in the header row
            # Find manager name: go backwards to find the manager row
            manager_name = None
            for scan_row in range(row_idx - 1, max(row_idx - 30, 0), -1):
                # Manager row has name in col base_col+1
                label_val = get_cell(scan_row, base_col + 1)
                if label_val and str(label_val).strip() in MANAGER_MAP:
                    manager_name = MANAGER_MAP[str(label_val).strip()]
                    break
                # Also check col base_col for some layouts
                label_val2 = get_cell(scan_row, base_col)
                if label_val2 and str(label_val2).strip() in MANAGER_MAP:
                    manager_name = MANAGER_MAP[str(label_val2).strip()]
                    break

            # Parse buyout entries below this header
            # Name col = base_col, Contract col = base_col+1, Salary col = base_col+2, FAAB col = base_col+3
            for b_row in range(row_idx + 1, row_idx + 10):
                b_name = get_cell(b_row, base_col)
                b_contract = get_cell(b_row, base_col + 1)
                b_salary_cost = get_cell(b_row, base_col + 2)
                b_faab_cost = get_cell(b_row, base_col + 3)

                # Stop at TOTAL or empty
                if b_name and str(b_name).strip() == "TOTAL":
                    break
                if not b_name and not b_contract:
                    continue

                player_name = str(b_name).strip() if b_name else ""
                contract_str = str(b_contract).strip() if b_contract else ""

                # Try to parse the contract string for buyout details
                parsed = parse_buyout_string(contract_str)

                # Also check if this is a special case like Wander Franco (Legal issue)
                note = ""
                # Check next columns for notes
                for note_col in [base_col + 3, base_col + 4]:
                    nv = get_cell(b_row, note_col)
                    if nv and "(Legal" in str(nv):
                        note = str(nv).strip()

                if player_name and (parsed or contract_str):
                    # Parse salary/faab costs safely
                    try:
                        sal_cost = abs(int(float(b_salary_cost))) if b_salary_cost and str(b_salary_cost).lstrip('-').replace('.', '').isdigit() else 0
                    except (ValueError, TypeError):
                        sal_cost = 0
                    try:
                        faab_cost_val = abs(int(float(b_faab_cost))) if b_faab_cost and str(b_faab_cost).lstrip('-').replace('.', '').isdigit() else 0
                    except (ValueError, TypeError):
                        faab_cost_val = 0

                    record = {
                        "team": manager_name or "UNKNOWN",
                        "player_name": player_name,
                        "contract_detail": contract_str,
                        "salary_cost": sal_cost,
                        "faab_cost": faab_cost_val,
                        "note": note,
                    }
                    if parsed:
                        record["original_contract"] = parsed["original_contract"]
                        record["parsed_salary"] = parsed["salary"]
                        record["parsed_contract_type"] = parsed["contract"]
                        record["parsed_faab_deduct"] = parsed["faab_cost"]
                        record["parsed_salary_deduct"] = parsed["salary_cost"]
                    buyout_records.append(record)

print("=" * 100)
print("  2025 PRE-DRAFT BUYOUT RECORDS")
print("=" * 100)
print()

for r in buyout_records:
    note_str = f" ({r['note']})" if r.get("note") else ""
    print(f"  {r['team']:12s} | {r['player_name']:25s} | {r['contract_detail']:20s} | "
          f"Salary Cap: -${r['salary_cost']:>3d} | FAAB: -${r['faab_cost']:>3d}{note_str}")

print()
print(f"Total buyout records: {len(buyout_records)}")

# Save for reference
with open("data/2025_buyout_records.json", "w") as f:
    json.dump(buyout_records, f, indent=2, ensure_ascii=False)
print(f"Saved to data/2025_buyout_records.json")
